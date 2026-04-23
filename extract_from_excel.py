"""
extract_from_excel.py
Extract μ, Σ, and prices from teammates' Excel files into CSVs
that the Streamlit app can read.

Place the two Excel files in the same directory, then run:
    python extract_from_excel.py

Outputs to data/ directory:
    - fund_info.csv
    - mu.csv               (monthly mean returns, matches Stats sheet)
    - mu_annualized.csv    (mu × 12, for display)
    - sigma.csv            (monthly covariance matrix)
    - sigma_annualized.csv (sigma × 12, for display)
    - prices.csv           (61 monthly price observations)
    - returns.csv          (60 monthly simple returns)
"""
import openpyxl
import pandas as pd
from pathlib import Path

# ---- File paths ----
EF_FILE = "EfficientFrontier_Part1.xlsm"
PRICES_FILE = "Monthly_Prices_61_Months_v2.xlsx"
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)


# ---- 1. Extract prices & fund names ----
wb_p = openpyxl.load_workbook(PRICES_FILE, data_only=True)
ws_p = wb_p["Monthly Prices"]

# Header row (row 1): ['Date', 'Franklin Technology', ...]
header = [c.value for c in ws_p[1]]
fund_names_full = header[1:]  # skip 'Date'

# Data rows
prices_data = []
for row in ws_p.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        break
    prices_data.append(row)

prices_df = pd.DataFrame(prices_data, columns=header)
prices_df = prices_df.rename(columns={"Date": "date"})
prices_df["date"] = pd.to_datetime(prices_df["date"])
prices_df = prices_df.set_index("date")

# Rename columns to Fund_01..Fund_10 to match Stats sheet
tickers = [f"Fund_{i:02d}" for i in range(1, 11)]
prices_df.columns = tickers
prices_df.to_csv(DATA_DIR / "prices.csv")

# ---- 2. Build fund_info.csv ----
# Asset class mapping (from README)
asset_classes = [
    ("Fund_01", "Franklin Technology A Acc SGD-H1", "Equity", "Global Tech"),
    ("Fund_02", "Allianz US Equity Cl AT Acc SGD", "Equity", "US"),
    ("Fund_03", "BlackRock World Healthscience A2 SGD-H", "Equity", "Global Healthcare"),
    ("Fund_04", "abrdn India Opportunities SGD", "Equity", "India"),
    ("Fund_05", "Amova Singapore Dividend Equity SGD", "Equity", "Singapore"),
    ("Fund_06", "Amova Japan Equity SGD", "Equity", "Japan"),
    ("Fund_07", "Allianz Oriental Income Cl AT Acc SGD", "Multi-Asset", "Asia"),
    ("Fund_08", "First Sentier Bridge A MDIS SGD", "Balanced", "Asia"),
    ("Fund_09", "United SGD Fund Cl A Acc SGD", "Bond", "SGD"),
    ("Fund_10", "LionGlobal Short Duration Bond A Acc SGD", "Bond", "SGD"),
]
fund_info = pd.DataFrame(asset_classes,
                         columns=["ticker", "name", "asset_class", "region"])
fund_info.to_csv(DATA_DIR / "fund_info.csv", index=False)


# ---- 3. Extract μ and Σ from Stats sheet ----
wb_ef = openpyxl.load_workbook(EF_FILE, data_only=True)
ws_s = wb_ef["Stats"]

# μ monthly: rows 4-13, column B (index 2)
mu_monthly = []
for row in range(4, 14):
    mu_monthly.append(ws_s.cell(row=row, column=2).value)

# Σ monthly: rows 19-28, columns B-K (2-11)
sigma_monthly = []
for row in range(19, 29):
    sigma_row = []
    for col in range(2, 12):
        sigma_row.append(ws_s.cell(row=row, column=col).value)
    sigma_monthly.append(sigma_row)

# Save monthly
mu_df = pd.DataFrame({"ticker": tickers, "expected_return": mu_monthly})
mu_df.to_csv(DATA_DIR / "mu.csv", index=False)

sigma_df = pd.DataFrame(sigma_monthly, index=tickers, columns=tickers)
sigma_df.to_csv(DATA_DIR / "sigma.csv")

# Save annualized (for display)
mu_ann = mu_df.copy()
mu_ann["expected_return"] = mu_ann["expected_return"] * 12
mu_ann.to_csv(DATA_DIR / "mu_annualized.csv", index=False)

sigma_ann = sigma_df * 12
sigma_ann.to_csv(DATA_DIR / "sigma_annualized.csv")


# ---- 4. Compute monthly returns for backtest ----
returns_df = prices_df.pct_change().dropna()
returns_df.to_csv(DATA_DIR / "returns.csv")


# ---- 5. Summary ----
print("✅ Extracted data to data/:")
for f in sorted(DATA_DIR.glob("*.csv")):
    print(f"  - {f.name}")

print("\n📊 Sanity checks:")
print(f"  Prices shape: {prices_df.shape}  (expected 61 × 10)")
print(f"  Returns shape: {returns_df.shape}  (expected 60 × 10)")
print(f"  μ (monthly): min={min(mu_monthly):.4f}, max={max(mu_monthly):.4f}")
print(f"  μ (annualized): min={min(mu_monthly)*12:.4f}, max={max(mu_monthly)*12:.4f}")
import numpy as np
eigvals = np.linalg.eigvalsh(np.array(sigma_monthly))
print(f"  Σ min eigenvalue: {eigvals.min():.6f}  (should be > 0 for PSD)")
